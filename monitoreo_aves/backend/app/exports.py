from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from math import log
import re
from statistics import mean
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session, joinedload

from . import database, models


router = APIRouter(prefix="/exports", tags=["exports"])

REPORT_VERSION = "1.0"
EXCEL_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

COLOR_DARK = "1F3328"
COLOR_GREEN = "2F6F4E"
COLOR_GREEN_LIGHT = "DDEADF"
COLOR_EARTH = "B88A2B"
COLOR_EARTH_LIGHT = "F5ECD7"
COLOR_RED = "B5483A"
COLOR_RED_LIGHT = "F7DEDA"
COLOR_BLUE_LIGHT = "DCEAF4"
COLOR_GREY = "E7ECE8"
COLOR_WHITE = "FFFFFF"
COLOR_TEXT = "243129"
COLOR_MUTED = "65736A"

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CED8D0"),
)
ILLEGAL_EXCEL_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

REVIEW_LABELS = {
    "unreviewed": "Sin revisar",
    "validated": "Validada",
    "corrected": "Corregida",
    "noise": "Ruido",
    "doubtful": "Dudosa",
    "discarded": "Descartada",
}

REVIEW_FILLS = {
    "Sin revisar": PatternFill("solid", fgColor=COLOR_GREY),
    "Validada": PatternFill("solid", fgColor=COLOR_GREEN_LIGHT),
    "Corregida": PatternFill("solid", fgColor=COLOR_BLUE_LIGHT),
    "Ruido": PatternFill("solid", fgColor=COLOR_EARTH_LIGHT),
    "Dudosa": PatternFill("solid", fgColor=COLOR_EARTH_LIGHT),
    "Descartada": PatternFill("solid", fgColor=COLOR_RED_LIGHT),
}


def _safe_text(value: object | None) -> str:
    """Limpia texto de usuario para que Excel no lo interprete como fórmula."""
    if value is None:
        return ""
    text = ILLEGAL_EXCEL_CHARACTERS.sub("", str(value))
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _clean_species_name(value: str | None) -> str:
    if not value:
        return "Desconocido"
    parts = value.split("_", 1)
    clean = parts[1] if len(parts) == 2 else parts[0]
    return clean.replace("_", " ").strip() or "Desconocido"


def _review_status(detection: models.Detection) -> str:
    return detection.review.status if detection.review else "unreviewed"


def _effective_species(detection: models.Detection) -> str:
    status = _review_status(detection)
    if (
        status == "corrected"
        and detection.review
        and detection.review.corrected_species
    ):
        return _clean_species_name(detection.review.corrected_species)
    if status == "noise":
        return "Ruido ambiente"
    if status == "discarded":
        return "Registro descartado"
    return _clean_species_name(detection.species)


def _is_valid_bird_detection(detection: models.Detection) -> bool:
    if _review_status(detection) in {"noise", "doubtful", "discarded"}:
        return False

    species = _effective_species(detection).casefold()
    excluded_terms = ("noise", "ruido", "human", "motor", "ambiente")
    return not any(term in species for term in excluded_terms)


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    return value.replace(tzinfo=None) if value.tzinfo is not None else value


def _date_range(
    date_from: date | None,
    date_to: date | None,
) -> tuple[datetime | None, datetime | None]:
    if date_from and date_to and date_from > date_to:
        raise HTTPException(
            status_code=422,
            detail="date_from no puede ser posterior a date_to",
        )

    start = datetime.combine(date_from, time.min) if date_from else None
    end = (
        datetime.combine(date_to + timedelta(days=1), time.min)
        if date_to
        else None
    )
    return start, end


def _apply_filters(query, model, start, end, device_id):
    if start is not None:
        query = query.filter(model.timestamp >= start)
    if end is not None:
        query = query.filter(model.timestamp < end)
    if device_id is not None:
        query = query.filter(model.device_id == device_id)
    return query


def _style_title(
    worksheet,
    title: str,
    subtitle: str,
    last_column: int,
) -> None:
    last_letter = get_column_letter(max(2, last_column))
    worksheet.merge_cells(f"A1:{last_letter}1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(
        name="Aptos Display",
        size=20,
        bold=True,
        color=COLOR_WHITE,
    )
    worksheet["A1"].fill = PatternFill("solid", fgColor=COLOR_DARK)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells(f"A2:{last_letter}2")
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(size=10, italic=True, color=COLOR_MUTED)
    worksheet["A2"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[2].height = 23


def _style_header_row(worksheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[row].height = 29


def _set_column_widths(
    worksheet,
    headers: list[str],
    rows: list[list[object]],
    requested_widths: dict[int, float] | None = None,
) -> None:
    requested_widths = requested_widths or {}
    for index, header in enumerate(headers, start=1):
        sample_values = [
            str(row[index - 1])
            for row in rows[:150]
            if index - 1 < len(row) and row[index - 1] is not None
        ]
        automatic = max([len(header), *(len(value) for value in sample_values)])
        width = requested_widths.get(index, min(max(automatic + 2, 10), 38))
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _create_table_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
    widths: dict[int, float] | None = None,
):
    worksheet = workbook.create_sheet(name)
    _style_title(worksheet, title, subtitle, len(headers))

    header_row = 4
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=column, value=header)
    _style_header_row(worksheet, header_row, len(headers))

    for row in rows:
        worksheet.append(
            [_safe_text(value) if isinstance(value, str) else value for value in row]
        )

    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        )
        worksheet.merge_cells(
            start_row=header_row + 2,
            start_column=1,
            end_row=header_row + 2,
            end_column=min(4, len(headers)),
        )
        empty_cell = worksheet.cell(row=header_row + 2, column=1)
        empty_cell.value = "No hay datos para los filtros seleccionados."
        empty_cell.font = Font(italic=True, color=COLOR_MUTED)

    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    _set_column_widths(worksheet, headers, rows, widths)
    return worksheet


def _apply_row_borders(worksheet, first_row: int, last_row: int, columns: int):
    for row in worksheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=1,
        max_col=columns,
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")


def _mean_attribute(
    items: list[models.AudioMetric],
    attribute: str,
) -> float | None:
    values = [
        float(value)
        for item in items
        if (value := getattr(item, attribute, None)) is not None
    ]
    return round(mean(values), 4) if values else None


def _biodiversity_indices(
    detections: list[models.Detection],
) -> tuple[int, int, float, float, float | None, str]:
    event_count = len(detections)
    counts = Counter(_effective_species(item) for item in detections)
    species_count = len(counts)

    if event_count == 0:
        return 0, 0, 0.0, 0.0, None, "SIN DATOS"

    proportions = [count / event_count for count in counts.values()]
    shannon = -sum(value * log(value) for value in proportions if value > 0)

    if event_count > 1:
        dominance = sum(count * (count - 1) for count in counts.values()) / (
            event_count * (event_count - 1)
        )
        simpson = 1 - dominance
    else:
        simpson = 0.0

    pielou = shannon / log(species_count) if species_count > 1 else None

    return (
        event_count,
        species_count,
        round(shannon, 3),
        round(simpson, 3),
        round(pielou, 3) if pielou is not None else None,
        "DESCRIPTIVO",
    )


def _build_detection_sheet(
    workbook: Workbook,
    detections: list[models.Detection],
    audio_by_event: dict[tuple[int, str], models.AudioMetric],
    base_url: str,
):
    headers = [
        "ID",
        "Timestamp",
        "Fecha",
        "Hora",
        "Nodo",
        "Ubicación",
        "Especie BirdNET",
        "Especie efectiva",
        "Confianza",
        "Revisión",
        "Especie corregida",
        "Revisor",
        "Nota",
        "Inicio BirdNET (s)",
        "Fin BirdNET (s)",
        "Amplitud RMS",
        "Calidad de captura",
        "Archivo WAV",
        "Enlace WAV",
    ]
    rows: list[list[object]] = []

    for detection in detections:
        timestamp = _excel_datetime(detection.timestamp)
        review = detection.review
        metric = audio_by_event.get((detection.device_id, detection.filename))
        rows.append(
            [
                detection.id,
                timestamp,
                timestamp.date() if timestamp else None,
                timestamp.time() if timestamp else None,
                detection.device.name if detection.device else detection.device_id,
                detection.device.location if detection.device else "Desconocida",
                _clean_species_name(detection.species),
                _effective_species(detection),
                detection.confidence,
                REVIEW_LABELS.get(_review_status(detection), _review_status(detection)),
                (
                    _clean_species_name(review.corrected_species)
                    if review and review.corrected_species
                    else ""
                ),
                review.reviewer if review else "",
                review.note if review else "",
                detection.audio_start_seconds,
                detection.audio_end_seconds,
                detection.amplitude,
                metric.quality_status if metric else "Sin telemetría",
                detection.filename,
                "Abrir WAV" if detection.filename else "",
            ]
        )

    worksheet = _create_table_sheet(
        workbook,
        "Detecciones",
        "Detecciones registradas",
        (
            "La especie efectiva incorpora las correcciones humanas sin modificar "
            "la clasificación original de BirdNET."
        ),
        headers,
        rows,
        "TablaDetecciones",
        {
            1: 9,
            2: 20,
            3: 12,
            4: 11,
            5: 21,
            6: 22,
            7: 27,
            8: 27,
            9: 12,
            10: 14,
            11: 27,
            12: 17,
            13: 36,
            14: 17,
            15: 17,
            16: 15,
            17: 19,
            18: 38,
            19: 15,
        },
    )

    if rows:
        last_row = 4 + len(rows)
        _apply_row_borders(worksheet, 5, last_row, len(headers))
        for row_number in range(5, last_row + 1):
            worksheet.cell(row_number, 2).number_format = "dd/mm/yyyy hh:mm:ss"
            worksheet.cell(row_number, 3).number_format = "dd/mm/yyyy"
            worksheet.cell(row_number, 4).number_format = "hh:mm:ss"
            worksheet.cell(row_number, 9).number_format = "0.0%"
            worksheet.cell(row_number, 14).number_format = "0.00"
            worksheet.cell(row_number, 15).number_format = "0.00"
            worksheet.cell(row_number, 16).number_format = "0.000000"

            review_cell = worksheet.cell(row_number, 10)
            review_cell.fill = REVIEW_FILLS.get(
                review_cell.value,
                PatternFill("solid", fgColor=COLOR_GREY),
            )

            quality_cell = worksheet.cell(row_number, 17)
            quality_status = str(quality_cell.value or "").casefold()
            quality_cell.fill = PatternFill(
                "solid",
                fgColor=(
                    COLOR_GREEN_LIGHT
                    if quality_status == "ok"
                    else COLOR_GREY
                    if quality_status in {"unknown", "sin telemetría"}
                    else COLOR_EARTH_LIGHT
                ),
            )

            filename = detections[row_number - 5].filename
            if filename:
                link_cell = worksheet.cell(row_number, 19)
                link_cell.hyperlink = (
                    f"{base_url}/records/{quote(filename, safe='')}"
                )
                link_cell.style = "Hyperlink"

        worksheet.conditional_formatting.add(
            f"I5:I{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=COLOR_RED,
                mid_type="num",
                mid_value=0.7,
                mid_color="F4D06F",
                end_type="num",
                end_value=1,
                end_color="63A46C",
            ),
        )
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:4"
    return worksheet


def _build_species_sheet(
    workbook: Workbook,
    valid_detections: list[models.Detection],
):
    grouped: dict[str, list[models.Detection]] = defaultdict(list)
    for detection in valid_detections:
        grouped[_effective_species(detection)].append(detection)

    rows = []
    for species, items in sorted(
        grouped.items(),
        key=lambda item: (-len(item[1]), item[0].casefold()),
    ):
        reviewed = sum(_review_status(item) != "unreviewed" for item in items)
        rows.append(
            [
                species,
                len(items),
                mean(item.confidence for item in items),
                reviewed,
                len(items) - reviewed,
            ]
        )

    worksheet = _create_table_sheet(
        workbook,
        "Especies",
        "Resumen por especie",
        (
            "Incluye detecciones válidas; excluye ruido, registros dudosos y "
            "descartados por revisión humana."
        ),
        [
            "Especie efectiva",
            "Detecciones",
            "Confianza media",
            "Revisadas",
            "Pendientes",
        ],
        rows,
        "TablaEspecies",
        {1: 32, 2: 15, 3: 18, 4: 14, 5: 14},
    )

    if rows:
        last_row = 4 + len(rows)
        _apply_row_borders(worksheet, 5, last_row, 5)
        for row_number in range(5, last_row + 1):
            worksheet.cell(row_number, 3).number_format = "0.0%"

        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Especies más detectadas"
        chart.y_axis.title = "Especie"
        chart.x_axis.title = "Detecciones"
        chart.height = 7.5
        chart.width = 12
        max_row = min(last_row, 14)
        chart.add_data(
            Reference(worksheet, min_col=2, min_row=4, max_row=max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=1, min_row=5, max_row=max_row)
        )
        chart.legend = None
        worksheet.add_chart(chart, "G4")

    return worksheet, rows


def _build_activity_sheet(
    workbook: Workbook,
    valid_detections: list[models.Detection],
):
    by_hour: dict[int, list[models.Detection]] = {
        hour: [] for hour in range(24)
    }
    for detection in valid_detections:
        by_hour[detection.timestamp.hour].append(detection)

    rows: list[list[object]] = []
    for hour in range(24):
        items = by_hour[hour]
        species = sorted({_effective_species(item) for item in items})
        rows.append(
            [
                f"{hour:02d}:00",
                f"{hour:02d}:00 - {hour:02d}:59",
                len(items),
                mean(item.confidence for item in items) if items else 0.0,
                len(species),
                ", ".join(species),
            ]
        )

    worksheet = _create_table_sheet(
        workbook,
        "Actividad horaria",
        "Actividad acústica por hora",
        (
            "Distribución agregada por hora del día para el periodo incluido "
            "en el informe."
        ),
        [
            "Hora",
            "Tramo horario",
            "Detecciones",
            "Confianza media",
            "Especies activas",
            "Taxones identificados",
        ],
        rows,
        "TablaActividadHoraria",
        {1: 11, 2: 19, 3: 15, 4: 18, 5: 18, 6: 48},
    )

    last_row = 4 + len(rows)
    _apply_row_borders(worksheet, 5, last_row, 6)
    for row_number in range(5, last_row + 1):
        worksheet.cell(row_number, 4).number_format = "0.0%"

    chart = LineChart()
    chart.style = 13
    chart.title = "Curva de actividad diaria"
    chart.y_axis.title = "Detecciones"
    chart.x_axis.title = "Hora"
    chart.height = 7.5
    chart.width = 13
    chart.add_data(
        Reference(worksheet, min_col=3, min_row=4, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(worksheet, min_col=1, min_row=5, max_row=last_row)
    )
    chart.legend = None
    worksheet.add_chart(chart, "H4")
    return worksheet, rows


def _build_biodiversity_sheet(
    workbook: Workbook,
    valid_detections: list[models.Detection],
    audio_metrics: list[models.AudioMetric],
):
    detections_by_device: dict[int, list[models.Detection]] = defaultdict(list)
    metrics_by_device: dict[int, list[models.AudioMetric]] = defaultdict(list)
    device_labels: dict[int, tuple[str, str]] = {}

    for detection in valid_detections:
        device_id = int(detection.device_id)
        node_name = (
            detection.device.name
            if detection.device and detection.device.name
            else f"Nodo {device_id}"
        )
        location = (
            detection.device.location
            if detection.device and detection.device.location
            else "Desconocida"
        )
        device_labels[device_id] = (node_name, location)
        detections_by_device[device_id].append(detection)

    for metric in audio_metrics:
        if metric.acoustic_metrics_version != "maad-v2":
            continue
        device_id = int(metric.device_id)
        node_name = (
            metric.device.name
            if metric.device and metric.device.name
            else f"Nodo {device_id}"
        )
        location = (
            metric.device.location
            if metric.device and metric.device.location
            else "Desconocida"
        )
        device_labels[device_id] = (node_name, location)
        metrics_by_device[device_id].append(metric)

    device_ids = sorted(
        set(detections_by_device) | set(metrics_by_device),
        key=lambda device_id: (
            device_labels[device_id][0].casefold(),
            device_id,
        ),
    )
    rows: list[list[object]] = []
    for device_id in device_ids:
        node_name, location = device_labels[device_id]
        event_count, species_count, shannon, simpson, pielou, scope = (
            _biodiversity_indices(detections_by_device[device_id])
        )
        metrics = metrics_by_device[device_id]
        rows.append(
            [
                node_name,
                location,
                scope,
                event_count,
                species_count,
                shannon,
                simpson,
                pielou,
                _mean_attribute(metrics, "rms"),
                _mean_attribute(metrics, "aci"),
                _mean_attribute(metrics, "adi"),
                _mean_attribute(metrics, "aei"),
                _mean_attribute(metrics, "bio"),
                _mean_attribute(metrics, "ndsi"),
                _mean_attribute(metrics, "ht"),
                _mean_attribute(metrics, "hf"),
                _mean_attribute(metrics, "h"),
            ]
        )

    headers = [
        "Nodo",
        "Zona",
        "Alcance de interpretación",
        "Eventos de detección N",
        "Especies detectadas S",
        "Shannon H'",
        "Simpson 1-D",
        "Pielou J'",
        "RMS medio",
        "ACI medio",
        "ADI medio",
        "AEI medio",
        "BIO medio",
        "NDSI medio",
        "Entropía temporal Ht",
        "Entropía frecuencial Hf",
        "Entropía acústica H",
    ]
    worksheet = _create_table_sheet(
        workbook,
        "Índices ecológicos",
        "Índices descriptivos de detecciones y del paisaje sonoro",
        (
            "Shannon, Simpson y Pielou describen el reparto de eventos válidos, "
            "no la abundancia de individuos. Los descriptores acústicos son "
            "medias de las capturas de cada nodo y solo deben compararse con el "
            "mismo nodo, micrófono, configuración y esfuerzo de muestreo. Solo "
            "se agrega la serie corregida maad-v2; el histórico legacy no se mezcla."
        ),
        headers,
        rows,
        "TablaIndicesEcologicos",
        {
            1: 23,
            2: 25,
            3: 25,
            4: 24,
            5: 23,
            **{column: 18 for column in range(6, 18)},
        },
    )

    if rows:
        last_row = 4 + len(rows)
        _apply_row_borders(worksheet, 5, last_row, len(headers))
        for row_number in range(5, last_row + 1):
            worksheet.cell(row_number, 3).fill = PatternFill(
                "solid", fgColor=COLOR_BLUE_LIGHT
            )
            for column in range(6, 18):
                worksheet.cell(row_number, column).number_format = "0.000"
    return worksheet


def _build_audio_quality_sheet(
    workbook: Workbook,
    audio_metrics: list[models.AudioMetric],
):
    headers = [
        "ID",
        "Timestamp",
        "Fecha",
        "Hora",
        "Nodo",
        "Ubicación",
        "Archivo WAV",
        "Muestreo (Hz)",
        "Duración (s)",
        "RMS",
        "Pico",
        "Clipping",
        "Offset DC",
        "Suelo de ruido RMS",
        "Estado",
        "Detalle",
        "Micrófono",
        "Modelo BirdNET",
        "Versión modelo",
        "Versión birdnetlib",
        "Versión métricas acústicas",
        "ACI",
        "ADI",
        "AEI",
        "BIO",
        "NDSI",
        "Ht",
        "Hf",
        "H",
    ]
    rows: list[list[object]] = []
    for metric in audio_metrics:
        timestamp = _excel_datetime(metric.timestamp)
        rows.append(
            [
                metric.id,
                timestamp,
                timestamp.date() if timestamp else None,
                timestamp.time() if timestamp else None,
                metric.device.name if metric.device else metric.device_id,
                metric.device.location if metric.device else "Desconocida",
                metric.filename,
                metric.sample_rate,
                metric.duration,
                metric.rms,
                metric.peak,
                metric.clipping_ratio,
                metric.dc_offset,
                metric.noise_floor_rms,
                metric.quality_status or "unknown",
                metric.quality_detail,
                metric.mic_device,
                metric.birdnet_model,
                metric.birdnet_model_version,
                metric.birdnetlib_version,
                metric.acoustic_metrics_version or "legacy-v1",
                metric.aci,
                metric.adi,
                metric.aei,
                metric.bio,
                metric.ndsi,
                metric.ht,
                metric.hf,
                metric.h,
            ]
        )

    worksheet = _create_table_sheet(
        workbook,
        "Calidad del audio",
        "Telemetría de las capturas",
        (
            "Una fila representa un ciclo de grabación, incluso cuando BirdNET "
            "no genera ninguna detección."
        ),
        headers,
        rows,
        "TablaCalidadAudio",
        {
            1: 9,
            2: 20,
            3: 12,
            4: 11,
            5: 22,
            6: 22,
            7: 38,
            15: 20,
            16: 42,
            17: 28,
            18: 25,
            19: 18,
            20: 20,
            21: 24,
        },
    )

    if rows:
        last_row = 4 + len(rows)
        _apply_row_borders(worksheet, 5, last_row, len(headers))
        for row_number in range(5, last_row + 1):
            worksheet.cell(row_number, 2).number_format = "dd/mm/yyyy hh:mm:ss"
            worksheet.cell(row_number, 3).number_format = "dd/mm/yyyy"
            worksheet.cell(row_number, 4).number_format = "hh:mm:ss"
            worksheet.cell(row_number, 9).number_format = "0.00"
            for column in (10, 11, 13, 14):
                worksheet.cell(row_number, column).number_format = "0.000000"
            worksheet.cell(row_number, 12).number_format = "0.000%"
            for column in range(22, 30):
                worksheet.cell(row_number, column).number_format = "0.000"

            status_cell = worksheet.cell(row_number, 15)
            status = str(status_cell.value or "").casefold()
            status_cell.fill = PatternFill(
                "solid",
                fgColor=(
                    COLOR_GREEN_LIGHT
                    if status == "ok"
                    else COLOR_GREY
                    if status == "unknown"
                    else COLOR_EARTH_LIGHT
                ),
            )
    return worksheet


def _build_reviews_sheet(
    workbook: Workbook,
    detections: list[models.Detection],
):
    reviewed = [
        detection
        for detection in detections
        if detection.review is not None
        and detection.review.status != "unreviewed"
    ]
    headers = [
        "ID detección",
        "Timestamp",
        "Nodo",
        "Ubicación",
        "Especie BirdNET",
        "Confianza BirdNET",
        "Resultado revisión",
        "Especie corregida",
        "Especie efectiva",
        "Revisor",
        "Nota",
        "Fecha de revisión",
        "Última actualización",
    ]
    rows: list[list[object]] = []
    for detection in reviewed:
        review = detection.review
        rows.append(
            [
                detection.id,
                _excel_datetime(detection.timestamp),
                detection.device.name if detection.device else detection.device_id,
                detection.device.location if detection.device else "Desconocida",
                _clean_species_name(detection.species),
                detection.confidence,
                REVIEW_LABELS.get(review.status, review.status),
                (
                    _clean_species_name(review.corrected_species)
                    if review.corrected_species
                    else ""
                ),
                _effective_species(detection),
                review.reviewer,
                review.note,
                _excel_datetime(review.reviewed_at),
                _excel_datetime(review.updated_at),
            ]
        )

    worksheet = _create_table_sheet(
        workbook,
        "Revisiones humanas",
        "Trazabilidad de la revisión humana",
        (
            "BirdNET se conserva como dato original; la revisión añade una capa "
            "de validación o corrección auditable."
        ),
        headers,
        rows,
        "TablaRevisionesHumanas",
        {
            1: 14,
            2: 20,
            3: 22,
            4: 22,
            5: 28,
            6: 20,
            7: 20,
            8: 28,
            9: 28,
            10: 18,
            11: 42,
            12: 20,
            13: 20,
        },
    )

    if rows:
        last_row = 4 + len(rows)
        _apply_row_borders(worksheet, 5, last_row, len(headers))
        for row_number in range(5, last_row + 1):
            worksheet.cell(row_number, 2).number_format = "dd/mm/yyyy hh:mm:ss"
            worksheet.cell(row_number, 6).number_format = "0.0%"
            worksheet.cell(row_number, 12).number_format = "dd/mm/yyyy hh:mm:ss"
            worksheet.cell(row_number, 13).number_format = "dd/mm/yyyy hh:mm:ss"
            status_cell = worksheet.cell(row_number, 7)
            status_cell.fill = REVIEW_FILLS.get(
                status_cell.value,
                PatternFill("solid", fgColor=COLOR_GREY),
            )
    return worksheet


def _format_period(
    date_from: date | None,
    date_to: date | None,
    timestamps: list[datetime],
) -> str:
    if date_from or date_to:
        start = date_from.isoformat() if date_from else "inicio"
        end = date_to.isoformat() if date_to else "actualidad"
        return f"{start} – {end}"
    if not timestamps:
        return "Sin datos"
    return (
        f"{min(timestamps).date().isoformat()} – "
        f"{max(timestamps).date().isoformat()}"
    )


def _build_metadata_sheet(
    workbook: Workbook,
    devices: list[models.Device],
    detections: list[models.Detection],
    audio_metrics: list[models.AudioMetric],
    date_from: date | None,
    date_to: date | None,
    device_id: int | None,
):
    worksheet = workbook.create_sheet("Metadatos")
    _style_title(
        worksheet,
        "Metadatos y metodología",
        "Información necesaria para interpretar y reproducir el informe.",
        4,
    )

    timestamps = [
        item.timestamp for item in [*detections, *audio_metrics] if item.timestamp
    ]
    model_names = sorted(
        {metric.birdnet_model for metric in audio_metrics if metric.birdnet_model}
    )
    model_versions = sorted(
        {
            metric.birdnet_model_version
            for metric in audio_metrics
            if metric.birdnet_model_version
        }
    )
    library_versions = sorted(
        {
            metric.birdnetlib_version
            for metric in audio_metrics
            if metric.birdnetlib_version
        }
    )
    selected_device = next(
        (device for device in devices if device.id == device_id),
        None,
    )

    metadata = [
        [
            "Versión del informe",
            REPORT_VERSION,
            "Versión de la estructura del libro Excel.",
        ],
        [
            "Generado en UTC",
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "Fecha y hora de generación del archivo.",
        ],
        [
            "Periodo incluido",
            _format_period(date_from, date_to, timestamps),
            "Rango inclusivo solicitado o rango real de los datos.",
        ],
        [
            "Filtro de nodo",
            (
                f"{selected_device.name} (ID {selected_device.id})"
                if selected_device
                else "Todos los nodos"
            ),
            "Nodo aplicado al informe.",
        ],
        [
            "Detecciones exportadas",
            len(detections),
            "Incluye detecciones válidas, ruido y registros revisados.",
        ],
        [
            "Ciclos de audio exportados",
            len(audio_metrics),
            "Filas de telemetría acústica, haya o no detección.",
        ],
        [
            "Modelo BirdNET",
            ", ".join(model_names) or "No registrado",
            "Nombre enviado por el nodo con la telemetría.",
        ],
        [
            "Versión del modelo",
            ", ".join(model_versions) or "No registrada",
            "Versión de BirdNET usada durante el periodo.",
        ],
        [
            "Versión de birdnetlib",
            ", ".join(library_versions) or "No registrada",
            "Versión de la biblioteca de inferencia.",
        ],
        [
            "Regla de especie efectiva",
            "Corrección humana > clasificación BirdNET",
            "La corrección se usa para agregados sin borrar el resultado original.",
        ],
        [
            "Regla de detección válida",
            "Excluye ruido, dudosas y descartadas",
            "Se aplica a especies, actividad e índices ecológicos.",
        ],
        [
            "Unidad de confianza",
            "Proporción 0–1 mostrada como porcentaje",
            "Puntuación proporcionada por BirdNET.",
        ],
    ]

    for column, header in enumerate(
        ["Parámetro", "Valor", "Descripción"],
        start=1,
    ):
        worksheet.cell(row=4, column=column, value=header)
    _style_header_row(worksheet, 4, 3)
    for row in metadata:
        worksheet.append(
            [
                _safe_text(value) if isinstance(value, str) else value
                for value in row
            ]
        )

    metadata_end = 4 + len(metadata)
    metadata_table = Table(
        displayName="TablaMetadatos",
        ref=f"A4:C{metadata_end}",
    )
    metadata_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showRowStripes=True,
    )
    worksheet.add_table(metadata_table)
    _apply_row_borders(worksheet, 5, metadata_end, 3)

    node_title_row = metadata_end + 3
    worksheet.merge_cells(
        start_row=node_title_row,
        start_column=1,
        end_row=node_title_row,
        end_column=6,
    )
    node_title = worksheet.cell(node_title_row, 1)
    node_title.value = "Nodos incluidos"
    node_title.font = Font(size=14, bold=True, color=COLOR_DARK)

    node_header_row = node_title_row + 1
    node_headers = [
        "ID",
        "Nodo",
        "Ubicación",
        "Coordenadas",
        "Origen ubicación",
        "Precisión declarada (m)",
    ]
    for column, header in enumerate(node_headers, start=1):
        worksheet.cell(node_header_row, column, value=header)
    _style_header_row(worksheet, node_header_row, len(node_headers))

    filtered_devices = (
        [device for device in devices if device.id == device_id]
        if device_id is not None
        else devices
    )
    for device in filtered_devices:
        coordinates = (
            f"{device.lat:.6f}, {device.lon:.6f}"
            if device.lat is not None and device.lon is not None
            else "No registradas"
        )
        worksheet.append(
            [
                device.id,
                _safe_text(device.name),
                _safe_text(device.location),
                coordinates,
                device.location_source or "unknown",
                device.location_accuracy_m,
            ]
        )

    node_end = node_header_row + len(filtered_devices)
    if filtered_devices:
        node_table = Table(
            displayName="TablaNodosInforme",
            ref=f"A{node_header_row}:F{node_end}",
        )
        node_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showRowStripes=True,
        )
        worksheet.add_table(node_table)
        _apply_row_borders(
            worksheet,
            node_header_row + 1,
            node_end,
            len(node_headers),
        )

    index_title_row = max(node_end, node_header_row) + 3
    worksheet.merge_cells(
        start_row=index_title_row,
        start_column=1,
        end_row=index_title_row,
        end_column=4,
    )
    index_title = worksheet.cell(index_title_row, 1)
    index_title.value = "Definición resumida de indicadores"
    index_title.font = Font(size=14, bold=True, color=COLOR_DARK)

    index_header_row = index_title_row + 1
    index_headers = ["Indicador", "Significado", "Interpretación resumida"]
    for column, header in enumerate(index_headers, start=1):
        worksheet.cell(index_header_row, column, value=header)
    _style_header_row(worksheet, index_header_row, 3)

    definitions = [
        [
            "Shannon H'",
            "Diversidad de detecciones",
            "Combina especies detectadas y reparto de eventos; no mide población.",
        ],
        [
            "Simpson 1-D",
            "Diversidad de detecciones",
            "Describe la concentración de eventos entre especies detectadas.",
        ],
        [
            "Pielou J'",
            "Equidad de detecciones",
            "Uniformidad del reparto de eventos entre especies detectadas.",
        ],
        ["ACI", "Complejidad acústica", "Variación temporal de energía por frecuencia."],
        ["ADI", "Diversidad acústica", "Distribución de energía entre bandas."],
        ["AEI", "Evenness acústico", "Desigualdad de ocupación entre bandas."],
        ["BIO", "Índice bioacústico", "Energía asociada a la banda biológica."],
        [
            "NDSI",
            "Balance espectral",
            "Balance de energía entre 0–1 kHz y 1–10 kHz; no clasifica por sí solo el entorno.",
        ],
        ["Ht / Hf / H", "Entropía acústica", "Heterogeneidad temporal, frecuencial y total."],
    ]
    for definition in definitions:
        worksheet.append([_safe_text(value) for value in definition])

    definitions_end = index_header_row + len(definitions)
    definitions_table = Table(
        displayName="TablaDefinicionesIndices",
        ref=f"A{index_header_row}:C{definitions_end}",
    )
    definitions_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showRowStripes=True,
    )
    worksheet.add_table(definitions_table)
    _apply_row_borders(worksheet, index_header_row + 1, definitions_end, 3)

    worksheet.column_dimensions["A"].width = 29
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 64
    worksheet.column_dimensions["D"].width = 24
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 24
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    return worksheet


def _build_summary_sheet(
    workbook: Workbook,
    detections: list[models.Detection],
    valid_detections: list[models.Detection],
    audio_metrics: list[models.AudioMetric],
    devices: list[models.Device],
    period: str,
    species_sheet,
    species_rows: list[list[object]],
    activity_sheet,
):
    worksheet = workbook.active
    worksheet.title = "Resumen"
    _style_title(
        worksheet,
        "BirdMonitor · Informe de monitorización",
        f"Resumen visual del periodo {period}",
        12,
    )
    worksheet.sheet_view.showGridLines = False

    reviewed = sum(_review_status(item) != "unreviewed" for item in detections)
    confidence = (
        mean(item.confidence for item in valid_detections)
        if valid_detections
        else 0.0
    )
    alerts = sum(
        str(metric.quality_status or "unknown").casefold()
        not in {"ok", "unknown"}
        for metric in audio_metrics
    )

    kpis = [
        ["Indicador", "Valor"],
        ["Detecciones totales", len(detections)],
        ["Detecciones válidas de aves", len(valid_detections)],
        ["Especies identificadas", len({_effective_species(item) for item in valid_detections})],
        ["Confianza media", confidence],
        ["Revisiones completadas", reviewed],
        ["Pendientes de revisión", len(detections) - reviewed],
        ["Ciclos de grabación", len(audio_metrics)],
        ["Alertas de calidad", alerts],
        ["Nodos incluidos", len(devices)],
    ]
    for row_index, row in enumerate(kpis, start=4):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row_index, column_index, value=value)

    _style_header_row(worksheet, 4, 2)
    for row_number in range(5, 4 + len(kpis)):
        worksheet.cell(row_number, 1).font = Font(bold=True, color=COLOR_TEXT)
        worksheet.cell(row_number, 1).fill = PatternFill(
            "solid",
            fgColor=COLOR_GREEN_LIGHT if row_number % 2 else "F3F6F4",
        )
        worksheet.cell(row_number, 2).font = Font(
            bold=True,
            size=13,
            color=COLOR_GREEN,
        )
        worksheet.cell(row_number, 2).alignment = Alignment(horizontal="center")
        worksheet.cell(row_number, 1).border = THIN_BORDER
        worksheet.cell(row_number, 2).border = THIN_BORDER
    worksheet.cell(8, 2).number_format = "0.0%"

    if species_rows:
        max_row = min(4 + len(species_rows), 14)
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Especies más detectadas"
        chart.y_axis.title = "Especie"
        chart.x_axis.title = "Detecciones"
        chart.height = 7.8
        chart.width = 13.5
        chart.add_data(
            Reference(species_sheet, min_col=2, min_row=4, max_row=max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(species_sheet, min_col=1, min_row=5, max_row=max_row)
        )
        chart.legend = None
        worksheet.add_chart(chart, "D4")
    else:
        worksheet["D5"] = "No hay detecciones válidas para representar."
        worksheet["D5"].font = Font(italic=True, color=COLOR_MUTED)

    activity_chart = LineChart()
    activity_chart.style = 13
    activity_chart.title = "Actividad por hora"
    activity_chart.y_axis.title = "Detecciones"
    activity_chart.x_axis.title = "Hora"
    activity_chart.height = 7.8
    activity_chart.width = 13.5
    activity_chart.add_data(
        Reference(activity_sheet, min_col=3, min_row=4, max_row=28),
        titles_from_data=True,
    )
    activity_chart.set_categories(
        Reference(activity_sheet, min_col=1, min_row=5, max_row=28)
    )
    activity_chart.legend = None
    worksheet.add_chart(activity_chart, "D20")

    worksheet["A17"] = "Cómo leer este informe"
    worksheet["A17"].font = Font(size=14, bold=True, color=COLOR_DARK)
    worksheet.merge_cells("A17:B17")
    worksheet["A18"] = (
        "Los agregados usan la especie corregida cuando existe. Las detecciones "
        "marcadas como ruido, dudosas o descartadas se conservan en Detecciones, "
        "pero no cuentan en actividad ni en los índices. N representa eventos de "
        "detección, no individuos. Los resultados son descriptivos y solo se "
        "comparan bajo el mismo nodo, micrófono, configuración y esfuerzo."
    )
    worksheet["A18"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet["A18"].font = Font(color=COLOR_MUTED)
    worksheet.merge_cells("A18:B22")

    worksheet.column_dimensions["A"].width = 31
    worksheet.column_dimensions["B"].width = 18
    for column in range(4, 13):
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    worksheet.freeze_panes = "A4"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = "A1:L36"
    return worksheet


def build_monitoring_workbook(
    detections: list[models.Detection],
    audio_metrics: list[models.AudioMetric],
    devices: list[models.Device],
    base_url: str,
    date_from: date | None = None,
    date_to: date | None = None,
    device_id: int | None = None,
) -> BytesIO:
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.creator = "BirdMonitor"
    workbook.properties.title = "Informe de monitorización de aves"
    workbook.properties.subject = "Detecciones, calidad acústica e índices ecológicos"
    workbook.properties.description = (
        "Informe Excel generado por BirdMonitor con trazabilidad de BirdNET "
        "y revisión humana."
    )
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)

    valid_detections = [
        detection
        for detection in detections
        if _is_valid_bird_detection(detection)
    ]
    audio_by_event = {
        (metric.device_id, metric.filename): metric for metric in audio_metrics
    }

    _build_detection_sheet(
        workbook,
        detections,
        audio_by_event,
        base_url.rstrip("/"),
    )
    species_sheet, species_rows = _build_species_sheet(
        workbook,
        valid_detections,
    )
    activity_sheet, _ = _build_activity_sheet(workbook, valid_detections)
    _build_biodiversity_sheet(workbook, valid_detections, audio_metrics)
    _build_audio_quality_sheet(workbook, audio_metrics)
    _build_reviews_sheet(workbook, detections)
    _build_metadata_sheet(
        workbook,
        devices,
        detections,
        audio_metrics,
        date_from,
        date_to,
        device_id,
    )

    period = _format_period(
        date_from,
        date_to,
        [
            item.timestamp
            for item in [*detections, *audio_metrics]
            if item.timestamp
        ],
    )
    _build_summary_sheet(
        workbook,
        detections,
        valid_detections,
        audio_metrics,
        devices,
        period,
        species_sheet,
        species_rows,
        activity_sheet,
    )

    workbook.active = workbook.sheetnames.index("Resumen")
    workbook["Resumen"].sheet_view.tabSelected = True
    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.tabColor = (
            COLOR_GREEN if worksheet.title != "Metadatos" else COLOR_EARTH
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("/report.xlsx")
def download_excel_report(
    request: Request,
    date_from: date | None = None,
    date_to: date | None = None,
    device_id: int | None = None,
    db: Session = Depends(database.get_db),
):
    start, end = _date_range(date_from, date_to)

    detection_query = (
        db.query(models.Detection)
        .options(
            joinedload(models.Detection.device),
            joinedload(models.Detection.review),
        )
        .order_by(models.Detection.timestamp.asc())
    )
    detection_query = _apply_filters(
        detection_query,
        models.Detection,
        start,
        end,
        device_id,
    )
    detections = detection_query.all()

    metric_query = (
        db.query(models.AudioMetric)
        .options(joinedload(models.AudioMetric.device))
        .order_by(models.AudioMetric.timestamp.asc())
    )
    metric_query = _apply_filters(
        metric_query,
        models.AudioMetric,
        start,
        end,
        device_id,
    )
    audio_metrics = metric_query.all()

    devices_query = db.query(models.Device).order_by(models.Device.name.asc())
    if device_id is not None:
        devices_query = devices_query.filter(models.Device.id == device_id)
    devices = devices_query.all()

    output = build_monitoring_workbook(
        detections=detections,
        audio_metrics=audio_metrics,
        devices=devices,
        base_url=str(request.base_url),
        date_from=date_from,
        date_to=date_to,
        device_id=device_id,
    )

    filename = f"birdmonitor_informe_{datetime.now().date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type=EXCEL_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )