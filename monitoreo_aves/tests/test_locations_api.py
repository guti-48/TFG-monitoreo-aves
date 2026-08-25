from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

def _activate(
    client,
    *,
    node,
    code,
    name,
    public_id,
    started_at,
    lat,
    lon,
):
    return client.post(
        "/node/deployments/activate",
        json={
            "device_name": node,
            "deployment_public_id": public_id,
            "site": {
                "code": code,
                "name": name,
                "municipality": name,
                "region": "Andalucia",
                "country_code": "ES",
                "lat": lat,
                "lon": lon,
                "location_source": "manual",
                "location_accuracy_m": 10.0,
                "timezone": "Europe/Madrid",
            },
            "started_at": started_at,
        },
    )


def _detection_payload(
    *,
    node,
    site_code,
    deployment_public_id,
    timestamp,
    species,
    filename,
):
    return {
        "species": species,
        "confidence": 0.9,
        "timestamp": timestamp,
        "filename": filename,
        "device_name": node,
        "site_code": site_code,
        "deployment_public_id": deployment_public_id,
        "amplitude": 0.02,
    }


def _metric_payload(
    *,
    node,
    site_code,
    deployment_public_id,
    timestamp,
    filename,
):
    return {
        "timestamp": timestamp,
        "filename": filename,
        "device_name": node,
        "site_code": site_code,
        "deployment_public_id": deployment_public_id,
        "sample_rate": 48000,
        "duration": 60.0,
        "rms": 0.01,
        "acoustic_metrics_version": "maad-v2",
        "aci": 1.0,
        "adi": 0.5,
        "aei": 0.4,
        "bio": 2.0,
        "ndsi": 0.1,
        "ht": 0.8,
        "hf": 0.7,
        "h": 0.56,
    }


def test_administra_sitios_sin_permitir_cambiar_su_codigo(client):
    created = client.post(
        "/sites/",
        json={
            "code": "fase3-sitio-manual",
            "name": "Sitio manual fase 3",
            "country_code": "es",
            "lat": 36.13,
            "lon": -5.45,
            "location_source": "manual",
        },
    )
    assert created.status_code == 200
    site = created.json()
    assert site["country_code"] == "ES"
    assert site["deployment_count"] == 0
    assert site["detection_count"] == 0

    duplicate = client.post(
        "/sites/",
        json={"code": "fase3-sitio-manual", "name": "Duplicado"},
    )
    assert duplicate.status_code == 409

    immutable_code = client.patch(
        f"/sites/{site['id']}",
        json={"code": "codigo-alterado"},
    )
    assert immutable_code.status_code == 422

    archived = client.patch(
        f"/sites/{site['id']}",
        json={"archived": True},
    )
    assert archived.status_code == 200
    assert archived.json()["archived_at"] is not None

    default_list = client.get("/sites/").json()
    assert site["id"] not in {item["id"] for item in default_list}
    full_list = client.get("/sites/", params={"include_archived": True}).json()
    assert site["id"] in {item["id"] for item in full_list}


def test_activacion_es_idempotente_y_cierra_el_despliegue_anterior(client):
    node = "fase3-node-activation"
    sevilla_uuid = "31000000-0000-4000-8000-000000000001"
    algeciras_uuid = "31000000-0000-4000-8000-000000000002"
    sevilla = _activate(
        client,
        node=node,
        code="fase3-sevilla-activation",
        name="Sevilla fase 3",
        public_id=sevilla_uuid,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.3845,
        lon=-6.0001,
    )
    assert sevilla.status_code == 200
    first = sevilla.json()
    assert first["active"] is True

    repeated = _activate(
        client,
        node=node,
        code="fase3-sevilla-activation",
        name="Sevilla fase 3",
        public_id=sevilla_uuid,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.3845,
        lon=-6.0001,
    )
    assert repeated.status_code == 200
    assert repeated.json()["id"] == first["id"]

    algeciras = _activate(
        client,
        node=node,
        code="fase3-algeciras-activation",
        name="Algeciras fase 3",
        public_id=algeciras_uuid,
        started_at="2026-08-10T00:00:00+00:00",
        lat=36.1408,
        lon=-5.4562,
    )
    assert algeciras.status_code == 200
    assert algeciras.json()["active"] is True

    delayed_replay = _activate(
        client,
        node=node,
        code="fase3-sevilla-activation",
        name="Sevilla fase 3",
        public_id=sevilla_uuid,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.3845,
        lon=-6.0001,
    )
    assert delayed_replay.status_code == 200
    assert delayed_replay.json()["active"] is False

    history = client.get(
        f"/devices/{first['device_id']}/deployments"
    ).json()
    assert len(history) == 2
    previous = next(item for item in history if item["public_id"] == sevilla_uuid)
    assert previous["active"] is False
    assert previous["ended_at"].startswith("2026-08-10T00:00:00")

    devices = client.get("/devices/").json()
    device = next(item for item in devices if item["id"] == first["device_id"])
    assert device["location"] == "Algeciras fase 3"


def test_ingesta_y_consultas_quedan_aisladas_por_sitio(client):
    node = "fase3-node-isolation"
    sevilla_uuid = "32000000-0000-4000-8000-000000000001"
    algeciras_uuid = "32000000-0000-4000-8000-000000000002"
    sevilla = _activate(
        client,
        node=node,
        code="fase3-sevilla-isolation",
        name="Sevilla aislamiento",
        public_id=sevilla_uuid,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.38,
        lon=-6.00,
    ).json()
    algeciras = _activate(
        client,
        node=node,
        code="fase3-algeciras-isolation",
        name="Algeciras aislamiento",
        public_id=algeciras_uuid,
        started_at="2026-08-10T00:00:00+00:00",
        lat=36.14,
        lon=-5.45,
    ).json()

    sevilla_detection = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-sevilla-isolation",
            deployment_public_id=sevilla_uuid,
            timestamp="2026-08-05T07:00:00+00:00",
            species="Species Sevilla Phase3",
            filename="phase3_sevilla.wav",
        ),
    )
    algeciras_detection = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-algeciras-isolation",
            deployment_public_id=algeciras_uuid,
            timestamp="2026-08-11T07:00:00+00:00",
            species="Species Algeciras Phase3",
            filename="phase3_algeciras.wav",
        ),
    )
    assert sevilla_detection.status_code == 200
    assert algeciras_detection.status_code == 200
    assert sevilla_detection.json()["site_id"] == sevilla["site_id"]
    assert algeciras_detection.json()["site_id"] == algeciras["site_id"]

    for site_code, public_id, timestamp, filename in (
        (
            "fase3-sevilla-isolation",
            sevilla_uuid,
            "2026-08-05T07:00:00+00:00",
            "phase3_sevilla.wav",
        ),
        (
            "fase3-algeciras-isolation",
            algeciras_uuid,
            "2026-08-11T07:00:00+00:00",
            "phase3_algeciras.wav",
        ),
    ):
        response = client.post(
            "/audio-metrics/",
            json=_metric_payload(
                node=node,
                site_code=site_code,
                deployment_public_id=public_id,
                timestamp=timestamp,
                filename=filename,
            ),
        )
        assert response.status_code == 200

    sevilla_rows = client.get(
        "/detections/",
        params={"site_id": sevilla["site_id"]},
    ).json()
    algeciras_rows = client.get(
        "/detections/",
        params={"site_id": algeciras["site_id"]},
    ).json()
    assert {row["species"] for row in sevilla_rows} == {
        "Species Sevilla Phase3"
    }
    assert {row["species"] for row in algeciras_rows} == {
        "Species Algeciras Phase3"
    }

    sevilla_metrics = client.get(
        "/audio-metrics/",
        params={"site_id": sevilla["site_id"]},
    ).json()
    assert {row["site_code"] for row in sevilla_metrics} == {
        "fase3-sevilla-isolation"
    }

    mismatch = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-sevilla-isolation",
            deployment_public_id=algeciras_uuid,
            timestamp="2026-08-12T07:00:00+00:00",
            species="Context mismatch",
            filename="phase3_mismatch.wav",
        ),
    )
    assert mismatch.status_code == 409

    partial_context = _detection_payload(
        node=node,
        site_code="fase3-sevilla-isolation",
        deployment_public_id=sevilla_uuid,
        timestamp="2026-08-06T07:00:00+00:00",
        species="Partial context",
        filename="phase3_partial.wav",
    )
    partial_context.pop("deployment_public_id")
    assert client.post("/detections/", json=partial_context).status_code == 422

    report = client.get(
        "/analytics/biodiversity",
        params={"site_id": sevilla["site_id"]},
    ).json()
    assert len(report) == 1
    assert report[0]["site_code"] == "fase3-sevilla-isolation"
    assert report[0]["riqueza"] == 1

    daily = client.get(
        "/analytics/daily-activity",
        params={"date": "2026-08-05", "site_id": sevilla["site_id"]},
    ).json()
    assert daily[7]["total_detecciones"] == 1

    map_data = client.get(
        "/analytics/map",
        params={"site_id": sevilla["site_id"]},
    ).json()
    assert map_data["site_code"] == "fase3-sevilla-isolation"
    assert map_data["event_count"] == 1

    mismatched_map = client.get(
        "/analytics/map",
        params={
            "site_id": sevilla["site_id"],
            "deployment_id": algeciras["id"],
        },
    ).json()
    assert mismatched_map["available"] is False
    assert "no pertenece" in mismatched_map["error"]

    excel = client.get(
        "/exports/report.xlsx",
        params={"site_id": sevilla["site_id"]},
    )
    assert excel.status_code == 200
    assert (
        'filename="birdmonitor_informe_fase3-sevilla-isolation_'
        in excel.headers["content-disposition"]
    )
    workbook = load_workbook(BytesIO(excel.content))
    exported_species = {
        row[6]
        for row in workbook["Detecciones"].iter_rows(
            min_row=5,
            values_only=True,
        )
        if row[0] is not None
    }
    assert "Species Sevilla Phase3" in exported_species
    assert "Species Algeciras Phase3" not in exported_species


def test_analisis_muestra_sitio_valido_aunque_tenga_cero_detecciones(client):
    node = "fase5-node-empty-site"
    public_id = "52000000-0000-4000-8000-000000000001"
    deployment = _activate(
        client,
        node=node,
        code="fase5-algeciras-empty",
        name="Algeciras fase 5 sin detecciones",
        public_id=public_id,
        started_at="2026-08-23T08:00:00+00:00",
        lat=36.12942,
        lon=-5.45303,
    ).json()

    metric = client.post(
        "/audio-metrics/",
        json=_metric_payload(
            node=node,
            site_code="fase5-algeciras-empty",
            deployment_public_id=public_id,
            timestamp="2026-08-23T08:01:00+00:00",
            filename="fase5_algeciras_metric.wav",
        ),
    )
    assert metric.status_code == 200
    assert client.get(
        "/detections/",
        params={"site_id": deployment["site_id"]},
    ).json() == []

    report_response = client.get(
        "/analytics/biodiversity",
        params={"site_id": deployment["site_id"]},
    )
    assert report_response.status_code == 200
    report = report_response.json()
    assert len(report) == 1
    assert report[0]["site_code"] == "fase5-algeciras-empty"
    assert report[0]["zona"] == "Algeciras fase 5 sin detecciones"
    assert report[0]["abundancia"] == 0
    assert report[0]["riqueza"] == 0
    assert report[0]["shannon"] == 0.0
    assert report[0]["simpson"] == 0.0
    assert report[0]["pielou"] is None
    assert report[0]["metrics_available"] is True
    assert report[0]["metric_samples"] == 1

    map_data = client.get(
        "/analytics/map",
        params={"site_id": deployment["site_id"]},
    ).json()
    assert map_data["site_code"] == "fase5-algeciras-empty"
    assert map_data["event_count"] == 0
    assert map_data["species_count"] == 0


def test_aprendizaje_de_una_especie_no_viaja_a_otro_sitio(client):
    node = "fase3-node-learning-site"
    site_a_uuid = "33000000-0000-4000-8000-000000000001"
    site_b_uuid = "33000000-0000-4000-8000-000000000002"
    site_a = _activate(
        client,
        node=node,
        code="fase3-learning-site-a",
        name="Sitio aprendizaje A",
        public_id=site_a_uuid,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.30,
        lon=-6.10,
    ).json()
    site_b = _activate(
        client,
        node=node,
        code="fase3-learning-site-b",
        name="Sitio aprendizaje B",
        public_id=site_b_uuid,
        started_at="2026-08-20T00:00:00+00:00",
        lat=36.20,
        lon=-5.40,
    ).json()

    for day in (2, 3, 4):
        detection = client.post(
            "/detections/",
            json=_detection_payload(
                node=node,
                site_code="fase3-learning-site-a",
                deployment_public_id=site_a_uuid,
                timestamp=f"2026-08-{day:02d}T08:00:00+00:00",
                species="Invasive Species Phase3",
                filename=f"phase3_learning_{day}.wav",
            ),
        ).json()
        reviewed = client.patch(
            f"/detections/{detection['id']}/review",
            json={"status": "validated", "reviewer": "pytest"},
        )
        assert reviewed.status_code == 200

    target_a = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-learning-site-a",
            deployment_public_id=site_a_uuid,
            timestamp="2026-08-05T08:00:00+00:00",
            species="Invasive Species Phase3",
            filename="phase3_learning_target_a.wav",
        ),
    ).json()
    target_b = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-learning-site-b",
            deployment_public_id=site_b_uuid,
            timestamp="2026-08-21T08:00:00+00:00",
            species="Invasive Species Phase3",
            filename="phase3_learning_target_b.wav",
        ),
    ).json()

    assert target_a["site_id"] == site_a["site_id"]
    assert target_a["learned_suggestion"] is not None
    assert target_b["site_id"] == site_b["site_id"]
    assert target_b["learned_suggestion"] is None

    site_a_rules = client.get(
        "/learning/rules",
        params={"site_id": site_a["site_id"], "active_only": True},
    ).json()
    site_b_rules = client.get(
        "/learning/rules",
        params={"site_id": site_b["site_id"], "active_only": True},
    ).json()
    assert any(
        rule["original_species"] == "Invasive Species Phase3"
        for rule in site_a_rules
    )
    assert not any(
        rule["original_species"] == "Invasive Species Phase3"
        for rule in site_b_rules
    )


def test_upload_se_almacena_y_recupera_dentro_de_su_despliegue(
    client,
    tmp_path,
    monkeypatch,
):
    from backend.app.features.detections import media as review_media
    from backend.app.features.uploads import routes as upload_routes

    records = tmp_path / "records"
    spectrograms = tmp_path / "spectrograms"
    records.mkdir()
    spectrograms.mkdir()
    monkeypatch.setattr(upload_routes, "SERVER_AUDIO_DIR", records)
    monkeypatch.setattr(upload_routes, "SPECTOGRAM_DIR", spectrograms)
    monkeypatch.setattr(review_media, "SERVER_AUDIO_DIR", records)
    monkeypatch.setattr(review_media, "SPECTOGRAM_DIR", spectrograms)

    node = "fase3-node-upload"
    public_id = "35000000-0000-4000-8000-000000000001"
    deployment = _activate(
        client,
        node=node,
        code="fase3-upload-site",
        name="Sitio subida fase 3",
        public_id=public_id,
        started_at="2026-08-01T00:00:00+00:00",
        lat=37.35,
        lon=-6.05,
    ).json()
    detection = client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-upload-site",
            deployment_public_id=public_id,
            timestamp="2026-08-02T09:00:00+00:00",
            species="Upload Species Phase3",
            filename="phase3_context_audio.wav",
        ),
    ).json()

    audio_bytes = b"RIFF-birdmonitor-phase3-context"
    spectrogram_bytes = b"\x89PNG\r\n\x1a\nphase3-context"
    upload = client.post(
        "/upload/",
        data={
            "deployment_public_id": public_id,
            "site_code": "fase3-upload-site",
            "device_name": node,
        },
        files={
            "audio": (
                "phase3_context_audio.wav",
                audio_bytes,
                "audio/wav",
            ),
            "specto": (
                "phase3_context_audio.png",
                spectrogram_bytes,
                "image/png",
            ),
        },
    )
    assert upload.status_code == 200
    assert upload.json()["deployment_id"] == deployment["id"]
    stored = (
        records
        / "fase3-upload-site"
        / public_id
        / "phase3_context_audio.wav"
    )
    assert stored.read_bytes() == audio_bytes
    stored_spectrogram = (
        spectrograms
        / "fase3-upload-site"
        / public_id
        / "phase3_context_audio.png"
    )
    assert stored_spectrogram.read_bytes() == spectrogram_bytes

    download = client.get(f"/detections/{detection['id']}/audio")
    assert download.status_code == 200
    assert download.content == audio_bytes
    assert download.headers["cache-control"] == "private, no-store"

    assert detection["spectrogram_url"] == (
        f"/detections/{detection['id']}/spectrogram"
    )
    spectrogram_download = client.get(detection["spectrogram_url"])
    assert spectrogram_download.status_code == 200
    assert spectrogram_download.content == spectrogram_bytes
    assert spectrogram_download.headers["content-type"] == "image/png"
    assert spectrogram_download.headers["cache-control"] == (
        "private, max-age=86400"
    )

    second_public_id = "35000000-0000-4000-8000-000000000002"
    _activate(
        client,
        node=node,
        code="fase3-upload-site-2",
        name="Segundo sitio de subida fase 3",
        public_id=second_public_id,
        started_at="2026-08-03T00:00:00+00:00",
        lat=36.15,
        lon=-5.45,
    )
    client.post(
        "/detections/",
        json=_detection_payload(
            node=node,
            site_code="fase3-upload-site-2",
            deployment_public_id=second_public_id,
            timestamp="2026-08-04T09:00:00+00:00",
            species="Second Upload Species Phase3",
            filename="phase3_context_audio.wav",
        ),
    )
    ambiguous_upload = client.post(
        "/upload/",
        files={
            "audio": (
                "phase3_context_audio.wav",
                audio_bytes,
                "audio/wav",
            )
        },
    )
    assert ambiguous_upload.status_code == 409
    assert "deployment_public_id" in ambiguous_upload.json()["detail"]


def test_payload_legacy_no_se_asigna_a_un_despliegue_cerrado(client):
    legacy = client.post(
        "/detections/",
        json={
            "species": "Legacy Species Phase3",
            "confidence": 0.8,
            "timestamp": "2026-08-01T08:00:00+00:00",
            "filename": "phase3_legacy_before_move.wav",
            "device_name": "birdmonitor",
            "amplitude": 0.01,
        },
    )
    assert legacy.status_code == 200
    assert legacy.json()["site_code"] == "sevilla"

    moved = _activate(
        client,
        node="birdmonitor",
        code="fase3-algeciras-legacy-cutoff",
        name="Algeciras tras payload legacy",
        public_id="36000000-0000-4000-8000-000000000001",
        started_at="2026-08-10T00:00:00+00:00",
        lat=36.14,
        lon=-5.45,
    )
    assert moved.status_code == 200

    late_legacy = client.post(
        "/detections/",
        json={
            "species": "Legacy Species Misplaced Phase3",
            "confidence": 0.8,
            "timestamp": "2026-08-11T08:00:00+00:00",
            "filename": "phase3_legacy_after_move.wav",
            "device_name": "birdmonitor",
            "amplitude": 0.01,
        },
    )
    assert late_legacy.status_code == 409
    assert "deployment_public_id" in late_legacy.json()["detail"]
