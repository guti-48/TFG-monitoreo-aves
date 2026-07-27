from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook


def test_informe_excel_contiene_datos_graficos_y_trazabilidad(client):
    device_response = client.post(
        "/devices/",
        json={
            "name": "raspberry-test-excel",
            "location": "Humedal Export Test",
            "lat": 37.3891,
            "lon": -5.9845,
        },
    )
    assert device_response.status_code == 200

    timestamp = datetime(2026, 7, 20, 6, 15, 0, tzinfo=timezone.utc)
    detection_response = client.post(
        "/detections/",
        json={
            "species": "Psittacula krameri_Rose-ringed Parakeet",
            "confidence": 0.87,
            "timestamp": timestamp.isoformat(),
            "filename": "excel_report_test.wav",
            "device_name": "raspberry-test-excel",
            "amplitude": 0.014,
            "audio_start_seconds": 12.0,
            "audio_end_seconds": 15.0,
        },
    )
    assert detection_response.status_code == 200
    detection = detection_response.json()

    review_response = client.patch(
        f"/detections/{detection['id']}/review",
        json={
            "status": "corrected",
            "corrected_species": "Rose-ringed Parakeet",
            "reviewer": "pytest",
            "note": "Confirmada mediante escucha.",
        },
    )
    assert review_response.status_code == 200

    metric_response = client.post(
        "/audio-metrics/",
        json={
            "timestamp": timestamp.isoformat(),
            "filename": "excel_report_test.wav",
            "device_name": "raspberry-test-excel",
            "sample_rate": 48000,
            "duration": 60.0,
            "rms": 0.014,
            "peak": 0.36,
            "clipping_ratio": 0.0001,
            "dc_offset": 0.0002,
            "noise_floor_rms": 0.004,
            "quality_status": "ok",
            "quality_detail": "Captura correcta.",
            "mic_device": "USB Audio Test",
            "birdnet_model": "BirdNET-Analyzer",
            "birdnet_model_version": "2.4",
            "birdnetlib_version": "0.18.1",
            "acoustic_metrics_version": "maad-v2",
            "aci": 110.2,
            "adi": 0.58,
            "aei": 0.72,
            "bio": 12.4,
            "ndsi": 0.22,
            "ht": 0.91,
            "hf": 0.63,
            "h": 0.57,
        },
    )
    assert metric_response.status_code == 200

    response = client.get(
        "/exports/report.xlsx",
        params={
            "date_from": "2026-07-20",
            "date_to": "2026-07-20",
            "device_id": detection["device_id"],
        },
    )

    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "Resumen",
        "Detecciones",
        "Especies",
        "Actividad horaria",
        "Índices ecológicos",
        "Calidad del audio",
        "Revisiones humanas",
        "Metadatos",
    ]

    assert workbook["Resumen"]["A1"].value == (
        "BirdMonitor · Informe de monitorización"
    )
    assert len(workbook["Resumen"]._charts) == 2

    ecological_sheet = workbook["Índices ecológicos"]
    ecological_headers = [
        cell.value for cell in ecological_sheet[4]
    ]
    assert ecological_headers[:5] == [
        "Nodo",
        "Zona",
        "Alcance de interpretación",
        "Eventos de detección N",
        "Especies detectadas S",
    ]
    ecological_rows = list(
        ecological_sheet.iter_rows(min_row=5, values_only=True)
    )
    exported_indices = next(
        row for row in ecological_rows if row[0] == "raspberry-test-excel"
    )
    assert exported_indices[1] == "Humedal Export Test"
    assert exported_indices[2] == "DESCRIPTIVO"
    assert exported_indices[3] == 1
    assert exported_indices[4] == 1
    assert all(
        value not in {"POBRE", "MODERADO", "EXCELENTE"}
        for row in ecological_rows
        for value in row
    )

    detections_sheet = workbook["Detecciones"]
    detection_rows = list(
        detections_sheet.iter_rows(min_row=5, values_only=False)
    )
    exported_row = next(
        row for row in detection_rows if row[0].value == detection["id"]
    )
    assert exported_row[7].value == "Rose-ringed Parakeet"
    assert exported_row[8].value == 0.87
    assert exported_row[8].number_format == "0.0%"
    assert exported_row[9].value == "Corregida"
    assert exported_row[16].value == "ok"
    assert exported_row[18].hyperlink.target == (
        "http://testserver/records/excel_report_test.wav"
    )

    audio_sheet = workbook["Calidad del audio"]
    audio_rows = list(audio_sheet.iter_rows(min_row=5, values_only=True))
    exported_audio = next(
        row for row in audio_rows if row[6] == "excel_report_test.wav"
    )
    assert exported_audio[14] == "ok"
    assert exported_audio[18] == "2.4"
    assert exported_audio[20] == "maad-v2"

    review_sheet = workbook["Revisiones humanas"]
    review_rows = list(review_sheet.iter_rows(min_row=5, values_only=True))
    exported_review = next(
        row for row in review_rows if row[0] == detection["id"]
    )
    assert exported_review[6] == "Corregida"
    assert exported_review[7] == "Rose-ringed Parakeet"


def test_informe_excel_rechaza_un_rango_invertido(client):
    response = client.get(
        "/exports/report.xlsx",
        params={"date_from": "2026-07-21", "date_to": "2026-07-20"},
    )

    assert response.status_code == 422
    assert "date_from" in response.json()["detail"]


def test_excel_no_mezcla_nodos_con_la_misma_ubicacion(client):
    timestamp = datetime(2026, 7, 21, 7, 0, 0, tzinfo=timezone.utc)
    for suffix, species in (("a", "Species alpha"), ("b", "Species beta")):
        node = f"excel-zona-compartida-{suffix}"
        device = client.post(
            "/devices/",
            json={
                "name": node,
                "location": "Zona compartida",
                "lat": 37.38,
                "lon": -5.98,
                "location_source": "manual",
            },
        )
        assert device.status_code == 200
        detection = client.post(
            "/detections/",
            json={
                "species": species,
                "confidence": 0.9,
                "timestamp": timestamp.isoformat(),
                "filename": f"{node}.wav",
                "device_name": node,
                "amplitude": 0.01,
            },
        )
        assert detection.status_code == 200

    response = client.get(
        "/exports/report.xlsx",
        params={"date_from": "2026-07-21", "date_to": "2026-07-21"},
    )
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    rows = list(
        workbook["Índices ecológicos"].iter_rows(
            min_row=5,
            values_only=True,
        )
    )
    shared_rows = [
        row for row in rows if str(row[0]).startswith("excel-zona-compartida-")
    ]
    assert len(shared_rows) == 2
    assert {row[0] for row in shared_rows} == {
        "excel-zona-compartida-a",
        "excel-zona-compartida-b",
    }
    assert all(row[1] == "Zona compartida" for row in shared_rows)
    assert all(row[3] == 1 for row in shared_rows)
    assert all(row[8] is None for row in shared_rows)